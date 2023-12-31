import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
import html


EXCEL_SOURCE_FILE = "_MMA21cL.xlsx"
MAX_POINTS = 40
# Feedback is provided for questions with text answers, also the student may read his answer
QUESTIONS_WITH_TEXT_ANSWERS = [
    "Stellen Sie sich vor, ein grosser Onlineshop",
    "Das folgende Diagramm beschreibt eine Musikdatenbank",
]

# GLOBALS
TRUNCATE_COL_VALUES = 80
data_grades = []  # List to store the grades


def read_excel(file_name):
    return pd.read_excel(EXCEL_SOURCE_FILE)


def compute_note_value(total_points):
    return (5 / MAX_POINTS) * total_points + 1


def rename_points_columns(df):
    points_cols = [col for col in df.columns if "Points" in col]
    rename_dict = {}
    for idx, col in enumerate(points_cols, 1):
        new_col_name = col.replace("Points", f"({idx})")
        if len(new_col_name) > TRUNCATE_COL_VALUES:
            new_col_name = new_col_name[: (TRUNCATE_COL_VALUES - 3)] + "..."
        rename_dict[col] = new_col_name
    df.rename(columns=rename_dict, inplace=True)
    return df, rename_dict


def filter_columns(df, rename_dict):
    first_six_cols = df.iloc[:, :6].columns.tolist()
    cols_to_keep = first_six_cols + list(rename_dict.values())
    df = df[cols_to_keep].copy()
    df.drop(
        columns=["ID", "Start time", "Completion time"], inplace=True, errors="ignore"
    )
    return df


def swap_name_order(name):
    name_parts = name.split()
    surname = name_parts[-1]
    first_name = " ".join(name_parts[:-1])
    return f"{surname} {first_name}"


def save_row_as_excel(row, max_points):
    name = swap_name_order(row["Name"])
    output_filename = (
        f'{EXCEL_SOURCE_FILE.replace("_", "").replace(".xlsx", "")}_{name}.xlsx'
    )
    row_df = pd.DataFrame([row])
    transposed_row = row_df.transpose()
    transposed_row.to_excel(output_filename, header=False, index=True)

    book = load_workbook(output_filename)
    sheet = book.active
    format_excel_sheet(sheet, row, max_points, name)

    # Reopen the source file to check for columns with "Feedback"
    sheet.append([None, None, None])
    sheet.append([None, None, None])
    source_df = read_excel(EXCEL_SOURCE_FILE)

    all_matching_columns = []  # Store all matching columns for all questions

    for question in QUESTIONS_WITH_TEXT_ANSWERS:
        matching_columns = [
            col
            for col in source_df.columns
            if question in str(col) and not col.startswith("Points")
        ]
        all_matching_columns.extend(matching_columns)

    for col in all_matching_columns:
        question_value = source_df.at[row.name, col]
        if pd.notna(question_value):
            question_value = str(question_value)
            # Convert line breaks and spaces to make them readable in Excel
            question_value = (
                question_value.replace("&nbsp;", " ")
                .replace("<br>", "\n")
                .replace("</span>", "")
                .replace("<span>", "")
            )
            question_value = html.unescape(question_value)
            if col.startswith("Feedback"):
                col = "Feedback"
            sheet.append(
                [col, question_value]
            )  # Use col instead of question to display the actual column name
            sheet.cell(row=sheet.max_row, column=1).alignment = Alignment(
                vertical="top", wrap_text=True
            )
            sheet.cell(row=sheet.max_row, column=2).alignment = Alignment(
                vertical="top", wrap_text=True
            )

    book.save(output_filename)


def format_excel_sheet(sheet, row, max_points, name):
    sheet.append(["", ""])
    total_points_value = row["Total points"]
    grade = compute_note_value(total_points_value)
    data_grades.append([name, total_points_value, grade])

    sheet.append(["Note", grade])
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="left")
    for cell in sheet["A"]:
        cell.font = Font(bold=False)
    for cell in sheet[sheet.max_row]:
        cell.font = Font(bold=True)
    sheet.column_dimensions["A"].width = 300 / 6
    sheet.column_dimensions["B"].width = 300 / 6


def export_grades_sheet():
    df_export = pd.DataFrame(data_grades, columns=["Name", "Total points", "Note"])

    # Export the DataFrame to Excel
    output_filename = f"_{EXCEL_SOURCE_FILE}_Notenblatt.xlsx"
    df_export.to_excel(output_filename, index=False)

    # Reopen the saved Excel file using openpyxl
    book = load_workbook(output_filename)
    sheet = book.active

    # Append an empty row
    sheet.append([None, None, None])

    # Compute the average of the "Note" column and append it
    avg_note = df_export["Note"].mean()
    sheet.append(["Average", None, avg_note])

    # Save the changes
    book.save(output_filename)


def main():
    df = read_excel(EXCEL_SOURCE_FILE)
    df, rename_dict = rename_points_columns(df)
    df = filter_columns(df, rename_dict)
    for _, row in df.iterrows():
        save_row_as_excel(row, MAX_POINTS)
    export_grades_sheet()


if __name__ == "__main__":
    main()
