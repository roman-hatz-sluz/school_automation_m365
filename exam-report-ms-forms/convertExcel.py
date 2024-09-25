import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
import html
import argparse
import json
import pprint

# Setup argument parser
parser = argparse.ArgumentParser(description="Process Excel file for M324 exam")
parser.add_argument(
    "--excel_source_path",
    type=str,
    required=True,
    help="Path to the Excel source directory",
)
parser.add_argument(
    "--excel_source_file", type=str, required=True, help="Excel source file name"
)
parser.add_argument(
    "--max_points", type=int, required=True, help="Maximum points for the exam"
)

args = parser.parse_args()

EXCEL_SOURCE_PATH = args.excel_source_path
EXCEL_SOURCE_FILE = args.excel_source_file
MAX_POINTS = args.max_points
JSON_FILE = "temp.json"

# GLOBALS
TRUNCATE_COL_VALUES = 800
data_grades = []  # List to store the grades


def read_json(file_path):
    with open(file_path, "r", encoding="utf-8") as file:
        return json.load(file)


def extract_text_field_questions(json_data):
    return [
        question["questionCount"]
        for question in json_data
        if question["type"] == "Question.TextField"
    ]


def read_excel(file_name):
    return pd.read_excel(EXCEL_SOURCE_PATH + EXCEL_SOURCE_FILE)


def compute_note_value(total_points):
    return (5 / MAX_POINTS) * total_points + 1


def rename_points_columns(df):
    points_cols = [col for col in df.columns if "Punkte" in col]
    rename_dict = {}
    for idx, col in enumerate(points_cols, 1):
        new_col_name = col.replace("Punkte", f"({idx})")
        if len(new_col_name) > TRUNCATE_COL_VALUES:
            new_col_name = new_col_name[: (TRUNCATE_COL_VALUES - 3)] + "..."
        rename_dict[col] = new_col_name
    df.rename(columns=rename_dict, inplace=True)
    return df, rename_dict


def filter_columns(df, rename_dict):
    first_six_cols = df.iloc[:, :6].columns.tolist()
    cols_to_keep = first_six_cols + list(rename_dict.values())
    df = df[cols_to_keep].copy()
    df.drop(
        columns=["ID", "Start time", "Completion time"], inplace=True, errors="ignore"
    )
    return df


def swap_name_order(name):
    name_parts = name.split()
    surname = name_parts[-1]
    first_name = " ".join(name_parts[:-1])
    return f"{surname} {first_name}"


def save_row_as_excel(row, max_points, text_questions):

    name = swap_name_order(row["Name"])
    output_filename = f'{EXCEL_SOURCE_PATH}/responses/{EXCEL_SOURCE_FILE.replace("_", "").replace(".xlsx", "")}_{name}.xlsx'
    row_df = pd.DataFrame([row])
    transposed_row = row_df.transpose()
    transposed_row.to_excel(output_filename, header=False, index=True)

    book = load_workbook(output_filename)
    sheet = book.active
    format_excel_sheet(sheet, row, max_points, name)

    # Reopen the source file to check for columns with "Feedback"
    source_df = read_excel(EXCEL_SOURCE_FILE)

    all_matching_columns = []  # Store all matching columns for all text questions

    for question_number in text_questions:
        matching_columns = [
            col
            for col in source_df.columns
            if f"({question_number})" in str(col) and not col.startswith("Punkte")
        ]
        all_matching_columns.extend(matching_columns)

    for col in all_matching_columns:
        question_value = source_df.at[row.name, col]
        if pd.notna(question_value):
            question_value = str(question_value)
            # Convert line breaks and spaces to make them readable in Excel
            question_value = (
                question_value.replace("&nbsp;", " ")
                .replace("<br>", "\n")
                .replace("</span>", "")
                .replace("<span>", "")
            )
            question_value = html.unescape(question_value)
            if col.startswith("Feedback"):
                col = "Feedback"
            sheet.append(
                [col, question_value]
            )  # Use col instead of question to display the actual column name
            sheet.cell(row=sheet.max_row, column=1).alignment = Alignment(
                vertical="top", wrap_text=True
            )
            sheet.cell(row=sheet.max_row, column=2).alignment = Alignment(
                vertical="top", wrap_text=True
            )

    book.save(output_filename)


def format_excel_sheet(sheet, row, max_points, name):
    sheet.append(["", ""])
    total_points_value = row["Gesamtpunktzahl"]
    grade = compute_note_value(total_points_value)
    data_grades.append([name, total_points_value, grade])
    sheet.append(["", ""])

    sheet.append(["Note", grade])
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="left")
    for cell in sheet["A"]:
        cell.font = Font(bold=False)
    for cell in sheet[sheet.max_row]:
        cell.font = Font(bold=True)
    sheet.column_dimensions["A"].width = 300 / 6
    sheet.column_dimensions["B"].width = 300 / 6


def export_grades_sheet():
    df_export = pd.DataFrame(data_grades, columns=["Name", "Gesamtpunktzahl", "Note"])

    # Export the DataFrame to Excel
    output_filename = f"{EXCEL_SOURCE_PATH}{EXCEL_SOURCE_FILE}_Notenblatt.xlsx"
    df_export.to_excel(output_filename, index=False)

    # Reopen the saved Excel file using openpyxl
    book = load_workbook(output_filename)
    sheet = book.active

    # Append an empty row
    sheet.append([None, None, None])

    # Compute the average of the "Note" column and append it
    avg_note = df_export["Note"].mean()
    sheet.append(["Average", None, avg_note])

    # Save the changes
    book.save(output_filename)

def create_question_structure_from_excel(df):
    # Initialize the list to store question data and column mappings
    questions = []
    question_count = 0

    # Get the column mapping for the first metadata columns
    mapping = {
        "E-Mail": df.columns.get_loc("E-Mail") + 1,  # Adjusting for 1-based indexing in Excel
        "Name": df.columns.get_loc("Name") + 1,
        "Gesamtpunktzahl": df.columns.get_loc("Gesamtpunktzahl") + 1,
    }

    # Questions start at column 8, each question takes up 3 columns (Question, Points, Feedback)
    first_question_col = 8
    question_mapping = {}

    # Calculate question column positions and build question entries
    col_names = df.columns.tolist()

    for i in range(first_question_col - 1, len(col_names), 3):  # Iterate over columns in sets of 3 (Question, Points, Feedback)
        if i < len(col_names):  # Ensure the question title column exists
            question_title = col_names[i]  # First column is the question title
        else:
            continue  # Skip if question title column doesn't exist

        
        # Increment question count
        question_count += 1

       
        # Create a question entry
        question_entry = {
            "question_title": question_title,
            "points_col": i + 2,
            "question_count": question_count,
            "row_index": i + 1  # 1-based index for the question's row in Excel
        }

        # Append the question entry to the questions list
        questions.append(question_entry)

        # Map this question into the question mapping (for further usage if needed)
        question_mapping[question_count] = {
            "title_col": i + 1,  
            "points_col": i + 2,
            "feedback_col": i + 3  
        }

    return mapping, questions, question_mapping

def main():
    json_data = read_json(JSON_FILE)
    text_questions = extract_text_field_questions(json_data["questions"])

    df = read_excel(EXCEL_SOURCE_FILE)
     
   
    pp = pprint.PrettyPrinter(indent=2)
    pp.pprint(create_question_structure_from_excel(df))
    exit(1)
    df, rename_dict = rename_points_columns(df)
    df = filter_columns(df, rename_dict)

    for _, row in df.iterrows():
        save_row_as_excel(row, MAX_POINTS, text_questions)
    export_grades_sheet()


if __name__ == "__main__":
    main()
