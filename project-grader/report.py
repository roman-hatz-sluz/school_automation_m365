#" python3 ./srv/excel-to-pdf-reporter/report.py --excel_file_path='./srv/excel-to-pdf-reporter/M290.24-25.PA-Noten.xlsx' "
 
import openpyxl
import re
from fpdf import FPDF  
import os
from report_utils import getGradeText, normalize, parse_arguments

args = parse_arguments()
excel_file = args.excel_file_path
current_dir = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = "./_secret/grades"

workbook = openpyxl.load_workbook(excel_file, data_only=True)
sheet = workbook.worksheets[0]
headers = sheet[1]
maxPoints = sheet[2]


def load_worksheet(excel_file_path):
    workbook = openpyxl.load_workbook(excel_file_path, data_only=True)
    return workbook.active, workbook.active[1], workbook.active[2]

def initialize_pdf():
    pdf = FPDF()
    pdf.add_page()
    pdf.add_font("DejaVu", "", f"{current_dir}/fonts/DejaVuSans.ttf", uni=True)
    pdf.add_font("DejaVu", "B", f"{current_dir}/fonts/DejaVuSans-Bold.ttf", uni=True)
    return pdf

def add_header_to_pdf(pdf, group_number):
    pdf.set_font("DejaVu", size=12)
    pdf.multi_cell(0, 8, f"Bewertung Projektarbeit - Gruppe {group_number} - 01.2025", align="L")
    pdf.set_font("DejaVu", size=10)
    pdf.multi_cell(0, 8, "Modul 290 - Datenbanken abfragen und verändern - BBZW Sursee - Lehrperson R. Hatz", align="L")

def add_footer_to_pdf(pdf):
    pdf.ln(4)
    pdf.ln(4)
    pdf.set_font("DejaVu", size=8)
    entry_text = """Hinweise:  
    - Beilagen: report.html für Aufgabe 3.2, images.html für Aufgabe 4.1
    - Alle Abschnitte (fett gedruckt) zusammen ergeben die Gesamtpunktzahl.
    - Aufgabe 2.3: je 1.5 Punkte für 1 Erweiterung.
    - Aufgabe 5.3: 2 Punkte für Detail Page, je 1 Punkt für die Challenges.
    - Allgemeines: Abgabedokument (Inhaltsverzeichnis, Design), Zwischenabgaben, Termine 
    """
    pdf.multi_cell(0, 4, entry_text, align="L")

def save_pdf(pdf, group_number):
    pdf_name = f"{OUTPUT_DIR}/{group_number}-report.pdf"
    pdf.output(pdf_name)
    print(f"PDF file created: {pdf_name}")

# Iterate over each row in the worksheet, starting from the third row
for row_index, row in enumerate(sheet.iter_rows(min_row=3), start=2):
    if row[0].value is None:
        continue
    # Initialize a new PDF document
    pdf = initialize_pdf()
    add_header_to_pdf(pdf, row[0].value)
   
    counter = 0
    for cell in row:
        counter = counter + 1
        if counter == 1:
            continue
        cell_value = str(cell.value) if cell.value else ""
        cell_value = normalize(cell_value)
        if cell_value == "":
            cell_value = "0"
        header = headers[counter - 1].value if headers[counter - 1].value else ""
        header = normalize(str(header))
        max_points = maxPoints[counter - 1].value
        if header == "":
            continue
        entry = f"{header}: {cell_value} von {max_points}"
        if float(cell_value) > float(max_points)   :
            print(f"Error: Cell value {cell_value} is above max points {max_points} (Gruppe: {row[0].value}, header: {header})")
            exit(1)
        if re.search("^[a-zA-Z]", header):
            pdf.ln(4)
            pdf.set_font("DejaVu", "B", 10)
            entry = f"{header}: {cell_value} von {max_points}"
        else:
            pdf.set_font("DejaVu", "", 10)
            entry = f"Aufgabe {header}: {cell_value} ({max_points})"
        if header == "Total":
            pdf.set_font("DejaVu", "B", 12)
        if header == "Note":
            pdf.set_font("DejaVu", "B", 12)
            entry = f"{header}: {round(float(cell_value), 3)}"

        pdf.multi_cell(0, 8, entry, align="L")
        if header == "Note":
            pdf.set_font("DejaVu", "B", 10)
            pdf.multi_cell(0, 8, getGradeText(cell_value), align="L")

        # Check for comments and append if present
        if cell.comment:
            comment_text = normalize(cell.comment.text.strip())
            pdf.set_font("DejaVu", size=8)  # Smaller font for comments

            for part in comment_text.split("- "):
                part = re.sub(r"\s+", " ", part).strip()
                if part:
                    pdf.multi_cell(0, 4, f"- {part}", align="L")
                    pdf.ln(1)

            pdf.set_font("DejaVu", size=10)  # Reset font size to normal

    add_footer_to_pdf(pdf)
   
    save_pdf(pdf,row[0].value)
